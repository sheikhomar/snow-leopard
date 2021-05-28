import os

from typing import List
from datetime import datetime
from pathlib import Path

import click
import numpy as np
import pandas as pd

from openpyxl import load_workbook


# From: https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def get_sorted_feature_names(dataframe, feature_cols, category: str):
    df_sorted = (dataframe[feature_cols]
        .groupby(dataframe.category_name)
        .count()
        .transpose()[[category]]
        .sort_values(category, ascending=False)
        .reset_index()
        .rename(columns={'index': 'feature_name'})
    )
    df_filtered = df_sorted[df_sorted[category] > 0].copy()
    
    df_filtered['feature_name_with_count'] = df_filtered['feature_name'] + ' (' + df_filtered[category].astype(str) + ')'
    result_dict = dict(zip(df_filtered.feature_name, df_filtered.feature_name_with_count))

    return result_dict
    

def get_sorted_categories(dataframe: pd.DataFrame):
    pass


def split(file_path: str):
    df_data = pd.read_csv(file_path, dtype=str, index_col=0)
    fixed_cols = [
        'id', 'supplier_id', 'supplier_name', 'category_id', 'category_name', 'title',
        'model_name', 'description_short', 'description_middle', 'description_long',
        'summary_short', 'summary_long', 'warranty', 'is_limited', 'on_market', 'quality',
        'url_details', 'url_manual', 'url_pdf', 'created_at', 'updated_at', 'released_on',
        'end_of_life_on', 'ean', 'n_variants', 'countries'
    ]
    feature_cols = list(set(list(df_data.columns)) - set(fixed_cols))

    df_fill_per_category = df_data[feature_cols].groupby(df_data.category_name).count()
    df_fill_per_category['n_rows'] = df_data[['id', 'category_name']].groupby('category_name').count()
    df_fill_per_category.sort_values('n_rows', ascending=False, inplace=True)

    categories = df_fill_per_category.index.tolist()
    n_total = len(categories)
    for i, category in enumerate(categories):
        print(f'{str(i+1).rjust(3)} / {n_total}: Exporting data for {category}...')
        sorted_feature_dict = get_sorted_feature_names(df_data, feature_cols, category)
        export_cols = fixed_cols + list(sorted_feature_dict.keys())
        df_filtered = df_data[df_data.category_name == category]
        df_export = df_filtered[export_cols].copy()
        sheet_name = category.replace('/', '-')
        append_df_to_excel('data/ice-cat-office-products-by-category.xlsx', df_export, sheet_name=sheet_name)



@click.command(help="Split products per category.")
@click.option(
    "-f",
    "--file-path",
    type=click.STRING,
    required=True,
)
def main(file_path: str):
    split(file_path)


if __name__ == "__main__":
    main()  # pylint: disable=no-value-for-parameter
